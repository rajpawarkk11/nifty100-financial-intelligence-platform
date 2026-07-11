"""
Sprint 3 Day 17
Screener Excel Export
"""

import os

import pandas as pd

from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.styles import Font

from src.screener.engine import run_screener
from src.screener.presets import PRESETS


OUTPUT = "output/screener_output.xlsx"


EXPORT_COLUMNS = [

    "company_id",
    "year",

    "return_on_equity_pct",
    "net_profit_margin_pct",
    "operating_profit_margin_pct",

    "debt_to_equity",
    "interest_coverage",

    "free_cash_flow_cr",
    "cash_from_operations_cr",

    "revenue_cagr_3yr",
    "revenue_cagr_5yr",

    "pat_cagr_3yr",
    "pat_cagr_5yr",

    "eps_cagr_3yr",
    "eps_cagr_5yr",

    "market_cap_crore",
    "pe_ratio",
    "pb_ratio",
    "dividend_yield_pct",

    "composite_quality_score"

]


os.makedirs("output", exist_ok=True)


with pd.ExcelWriter(
    OUTPUT,
    engine="openpyxl"
) as writer:

    for preset_name, preset_filters in PRESETS.items():

        df = run_screener(preset_filters)

        columns = [
            c for c in EXPORT_COLUMNS
            if c in df.columns
        ]

        df = df[columns]

        df = df.sort_values(
            "composite_quality_score",
            ascending=False
        )

        df.to_excel(
            writer,
            sheet_name=preset_name[:31],
            index=False
        )


wb = load_workbook(OUTPUT)

GREEN = PatternFill(
    fill_type="solid",
    start_color="C6EFCE",
)

RED = PatternFill(
    fill_type="solid",
    start_color="FFC7CE",
)


def apply_threshold(cell, operator, threshold):
    """
    Apply green/red formatting based on threshold.
    """

    if cell.value is None:
        return

    try:
        value = float(cell.value)
    except (TypeError, ValueError):
        return

    passed = False

    if operator == ">=":
        passed = value >= threshold

    elif operator == "<=":
        passed = value <= threshold

    elif operator == "==":
        passed = value == threshold

    if passed:
        cell.fill = GREEN
    else:
        cell.fill = RED

for ws in wb.worksheets:

    preset_filters = PRESETS.get(ws.title)

    if preset_filters is None:
        continue

    headers = [
        cell.value
        for cell in ws[1]
    ]

    for row in ws.iter_rows(min_row=2):

        values = {
            headers[i]: row[i]
            for i in range(len(headers))
        }

        for rule in preset_filters.values():

            column = rule["column"]

            operator = rule["operator"]

            threshold = rule["value"]

            if column not in values:
                continue

            apply_threshold(
                values[column],
                operator,
                threshold
            )
wb.save(OUTPUT)

print("=" * 60)
print("Excel Generated Successfully")
print(OUTPUT)
print("=" * 60)