"""
Sprint 3 Day 20
Peer Comparison Excel Report
"""

import sqlite3
import os

import pandas as pd

from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font

DB_PATH = "db/nifty100.db"
OUTPUT_FILE = "output/peer_comparison.xlsx"

os.makedirs("output", exist_ok=True)


def load_data():

    conn = sqlite3.connect(DB_PATH)

    peer_percentiles = pd.read_sql(
        """
        SELECT *
        FROM peer_percentiles
        """,
        conn,
    )

    peer_groups = pd.read_sql(
        """
        SELECT *
        FROM peer_groups
        """,
        conn,
    )

    companies = pd.read_sql(
        """
        SELECT
            id,
            company_name
        FROM companies
        """,
        conn,
    )

    conn.close()

    return (
        peer_percentiles,
        peer_groups,
        companies,
    )


def build_peer_table(
    peer_percentiles,
    peer_groups,
    companies,
):

    values = peer_percentiles.pivot_table(

        index=[
            "company_id",
            "peer_group_name",
            "year",
        ],

        columns="metric",

        values="value",

        aggfunc="first",

    ).reset_index()

    percentiles = peer_percentiles.pivot_table(

        index=[
            "company_id",
            "peer_group_name",
            "year",
        ],

        columns="metric",

        values="percentile_rank",

        aggfunc="first",

    )

    percentiles.columns = [
        f"{c}_percentile"
        for c in percentiles.columns
    ]

    percentiles = percentiles.reset_index()

    df = values.merge(

        percentiles,

        on=[
            "company_id",
            "peer_group_name",
            "year",
        ],

    )

    df = df.merge(

        companies,

        left_on="company_id",
        right_on="id",
        how="left",

    )

    df = df.merge(

        peer_groups[
            [
                "company_id",
                "is_benchmark",
            ]
        ],

        on="company_id",
        how="left",

    )

    return df
def export_excel(df):

    with pd.ExcelWriter(
        OUTPUT_FILE,
        engine="openpyxl",
    ) as writer:

        peer_groups = sorted(

            df[
                df["peer_group_name"]
                != "No peer group assigned"
            ]["peer_group_name"].unique()

        )

        for peer in peer_groups:

            sheet = (

                df[
                    df["peer_group_name"]
                    == peer
                ]

                .sort_values(

                    "return_on_equity_pct_percentile",

                    ascending=False,

                )

            )

            sheet.to_excel(

                writer,

                sheet_name=peer[:31],

                index=False,

            )

    print("=" * 60)
    print("Peer Comparison Excel Generated")
    print(OUTPUT_FILE)
    print("=" * 60)
def apply_formatting():

    wb = load_workbook(
        OUTPUT_FILE
    )

    GREEN = PatternFill(
        fill_type="solid",
        start_color="C6EFCE",
    )

    YELLOW = PatternFill(
        fill_type="solid",
        start_color="FFF2CC",
    )

    RED = PatternFill(
        fill_type="solid",
        start_color="FFC7CE",
    )

    GOLD = PatternFill(
        fill_type="solid",
        start_color="FFD966",
    )

    for ws in wb.worksheets:

        headers = [
            c.value
            for c in ws[1]
        ]

        percentile_columns = [

            i

            for i, h in enumerate(headers)

            if (
                isinstance(h, str)
                and h.endswith("_percentile")
            )

        ]

        benchmark_index = None

        if "is_benchmark" in headers:

            benchmark_index = headers.index(
                "is_benchmark"
            )

        for row in ws.iter_rows(min_row=2):

            if benchmark_index is not None:

                value = row[
                    benchmark_index
                ].value

                if value == 1:

                    for cell in row:
                        cell.fill = GOLD
                        cell.font = Font(
                            bold=True
                        )

            for idx in percentile_columns:

                cell = row[idx]

                if cell.value is None:
                    continue

                try:
                    value = float(
                        cell.value
                    )
                except Exception:
                    continue

                if value >= 75:
                    cell.fill = GREEN

                elif value <= 25:
                    cell.fill = RED

                else:
                    cell.fill = YELLOW

    wb.save(
        OUTPUT_FILE
    )
def add_median_rows():

    wb = load_workbook(OUTPUT_FILE)

    for ws in wb.worksheets:

        max_row = ws.max_row
        max_col = ws.max_column

        headers = [
            cell.value
            for cell in ws[1]
        ]

        median_row = max_row + 2

        ws.cell(
            row=median_row,
            column=1
        ).value = "Peer Group Median"

        ws.cell(
            row=median_row,
            column=1
        ).font = Font(bold=True)

        for col in range(2, max_col + 1):

            values = []

            for r in range(2, max_row + 1):

                value = ws.cell(r, col).value

                if isinstance(
                    value,
                    (int, float)
                ):
                    values.append(value)

            if values:

                ws.cell(
                    row=median_row,
                    column=col
                ).value = round(
                    pd.Series(values).median(),
                    2
                )

    wb.save(OUTPUT_FILE)


if __name__ == "__main__":

    peer_percentiles, peer_groups, companies = load_data()

    df = build_peer_table(
        peer_percentiles,
        peer_groups,
        companies,
    )

    export_excel(df)

    apply_formatting()

    add_median_rows()

    print("=" * 60)
    print("Peer Comparison Report Completed")
    print("Output :", OUTPUT_FILE)
    print("Peer Groups :", df["peer_group_name"].nunique() - 1)
    print("=" * 60)